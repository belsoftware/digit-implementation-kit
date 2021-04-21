from common import *
from config import config, getValue
import io
import os 
import sys
from common import superuser_login
from PropertyTax import *
import pandas as pd
import openpyxl
import collections
import re
from CreateWaterConnection import *
from CreateSewerageConnection import *
import traceback


INDEX_SL_NO = 0
INDEX_ABAS_PROPERTY_ID = 1
INDEX_OLD_PROPERTY_ID = 2
INDEX_PROPERTY_TYPE = 3
INDEX_TOTAL_LAND_AREA = 4
INDEX_TOTAL_CONSTRUCTED_AREA = 5
INDEX_USAGE_TYPE = 6
INDEX_TENANT_REGIONNAME = 7
INDEX_TENANT_REGIONCODE = 8
INDEX_TENANT_LATITUDE = 9
INDEX_TENANT_LONGITUDE = 10
INDEX_TENANT_CONTACT = 12
INDEX_TENANT_EMAIL = 13
INDEX_TENANT_ADDRESS = 14
INDEX_TENANT_FB = 17
INDEX_TENANT_TWITTER = 18
INDEX_TENANT_POPULATION = 20
INDEX_TENANT_GRADE = "Cantonment Board"
INDEX_STATE = 30
INDEX_CITY_HINDI = 31
INDEX_DISTRICT_HINDI = 32
INDEX_STATE_HINDI = 33
FOLDER_PATH  =r'C:\Users\Administrator\Downloads\WaterSewerageTemplates'

def main() :
    print("Replace 109 of C:\ProgramData\Miniconda3\envs\py36\lib\site-packages\openpyxl\worksheet\merge.py with below one ") 
    print ("if side is None or  side.style is None:")
    root = FOLDER_PATH
    errorlogfile = open(os.path.join(root, "errorCBs.txt"), "w")  
    successlogfile = open(os.path.join(root, "CB With ProperData.txt"), "w")
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        for found_index, module in enumerate(cb_module_data["tenants"]):
            if module["city"]["ulbGrade"]=="ST":
                continue
            cityname =module["code"].lower()[3:]
            
            name = 'CB ' + cityname.lower()
            if  os.path.exists( os.path.join(root,name)):                
                try : 
                    if  True : # cityname =='clementtown' : #'roorkee'  :
                        print("Processing for CB "+cityname.upper())
                        config.CITY_NAME = cityname
                        cbMain(cityname, successlogfile)
                except Exception as ex: 
                    print("Error in processing CB ",cityname , ex)
                    traceback.print_exc()
                    errorlogfile.write(cityname+"\n")
    errorlogfile.close()
    successlogfile.close()


def cbMain(cityname, successlogfile):
    Flag =False
    tenantMapping={}
    count = 0
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        for found_index, module in enumerate(cb_module_data["tenants"]):
            if module["city"]["ulbGrade"]=="ST":
                continue
            tenantMapping[module["code"].lower()]=module["code"].lower()[3:]


    # Doing for one cb at a time
    # cityname = 'wellington'
    root = FOLDER_PATH
    name = 'CB ' + cityname.lower()
    propertyFile =os.path.join(root, name,'Template for Existing Property-Integrated with ABAS-' + cityname + '.xlsx')
    waterFile = os.path.join(root, name, "Template for Existing Water Connection Detail.xlsx")
    sewerageFile = os.path.join(root, name, "Template for Existing Sewerage Connection Detail.xlsx")
    logfile = open(os.path.join(root, name, "Logfile.json"), "w")   
    logfile.write("[ ")  
    if config.INSERT_DATA :
        validate = enterDefaultMobileNo(propertyFile, tenantMapping, cityname, waterFile, sewerageFile,logfile) 
        if(validate == False):                
            print('Data validation Failed for mobile entry, Please check the log file.') 
            return
    
    if os.path.exists(propertyFile) :  
        validate =  validateDataForProperty(propertyFile, logfile)            
        if(validate == False):                
            print('Data validation for property Failed, Please check the log file.') 
            if config.INSERT_DATA: 
                return
        else:
            print('Data validation for property success.')                
        wb_property = openpyxl.load_workbook(propertyFile) 
        sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')         
        sheet2 = wb_property.get_sheet_by_name('Property Ownership Details')
        
        localitySheet = wb_property.get_sheet_by_name('Locality')
        df = pd.read_excel(propertyFile, sheet_name='Locality', usecols=['Locality Name', 'Code'])
        locality_data = {}
        for ind in df.index: 
            locality_data[df['Locality Name'][ind]] =  df['Code'][ind]   
        if config.INSERT_DATA: 
            createPropertyJson(sheet1, sheet2, locality_data,cityname, logfile, root, name)                
            wb_property.save(propertyFile)        
        wb_property.close()
    else:
        print("Property File doesnot exist for ", cityname) 
    
    if os.path.exists(waterFile) : 
        ProcessWaterConnection(propertyFile, waterFile, logfile, root, name,  cityname)  
        
    else:
        print("Water File doesnot exist for ", cityname) 

    if os.path.exists(sewerageFile) : 
        ProcessSewerageConnection(propertyFile, sewerageFile, logfile, root, name,  cityname)  
    else:
        print("Sewerage File doesnot exist for ", cityname) 

    logfile.seek(logfile.tell() - 1, os.SEEK_SET)
    logfile.write('')
    logfile.write("]")        

    size = os.path.getsize(os.path.join(root, name, "Logfile.json")) 
    logfile.close() 
    try :       
        now = datetime.now()
        date_time = now.strftime("%d-%m-%Y") 
        if size > 2 : 
            df = pd.read_json (os.path.join(root, name, "Logfile.json"))
            if not os.path.exists(os.path.join(root,date_time + '-Data_Entries_Issues')) :
                os.makedirs(os.path.join(root,date_time + '-Data_Entries_Issues'))
            df.to_excel(os.path.join(root, date_time + '-Data_Entries_Issues',    name+ " Data Entries Issues.xlsx"), index = None)
            #df.to_excel(os.path.join(root, "WS_Data_Entry_Issues","CB "+ cityname+ " - Data Entries Issues.xlsx"), index = None)
            #df.to_csv (os.path.join(root, name, "DataValidation.csv"), index = None)
        else : 
            successlogfile.write("\n")
            successlogfile.write(cityname)
            successlogfile.write("\n")            
    except Exception as ex: 
        print("Error in parsing json file",ex)


def validateDataForProperty(propertyFile, logfile):
    validated = True
    reason = ''

    try:
        wb_property = openpyxl.load_workbook(propertyFile) 
        sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')   
        sheet2 = wb_property.get_sheet_by_name('Property Ownership Details')
        abas_ids = []        
        reason = 'Property file validation starts.\n'
        # validated = ValidateCols(logfile, propertyFile, sheet1, sheet2)
        #print('no. of rows in Property file sheet 1: ', sheet1.max_row ) 
        emptyRows=0
        count =0 
        for row in sheet1.iter_rows(min_row=3, max_col=42, max_row=sheet1.max_row ,values_only=True): 
            try : 
                if emptyRows > 10 :
                    break
                #print(row[0] ,"  ",row[1] ,"   ",row[2]   )
                if pd.isna(row[0]) and pd.isna(row[1])  and pd.isna(row[2]):
                    emptyRows =emptyRows +1
                    continue
                if pd.isna(row[0]):
                    validated = False
                    reason = 'Sl no. column is empty'
                    write(logfile,propertyFile,sheet1.title,row[0],reason)
                    #logfile.write(reason)
                    break
                if pd.isna(row[1]):
                    validated = False
                    #print("Property ID ",row[1])
                    reason = 'Property File sheet1 data validation failed for sl no. '+ str(row[0]) + ', abas property id  is empty.\n'
                    write(logfile,propertyFile,sheet1.title,row[0],'abas property id  is empty',row[1])
                    #logfile.write(reason)
                if pd.isna(row[27]):
                    validated = False
                    reason = 'Property File sheet1 data validation failed for sl no. '+ str(row[0]) + ',  ownership type is empty.\n'
                    write(logfile,propertyFile,sheet1.title,row[0],'ownership type is empty',row[1])
                    #logfile.write(reason)
                if pd.isna(row[13]):
                    validated = False
                    reason = 'Property File sheet1 data validation failed for sl no. '+ str(row[0]) + ',  Locality/ Mohalla is empty.\n'
                    write(logfile,propertyFile,sheet1.title,row[0],'Locality/ Mohalla is empty',row[1])
                    #logfile.write(reason)
                if(str(row[27]) != "Multiple Owners"):
                    if pd.isna(row[28]):
                        validated = False
                        reason = 'Property File sheet1 data validation failed for sl no. '+ str(row[0]) + ', name is empty.\n'
                        write(logfile,propertyFile,sheet1.title,row[0],'name is empty',row[1])
                        #logfile.write(reason)
                    # elif not pd.isna(row[28]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[28]))):                    
                    #     validated = False
                    #     reason = 'Name has invalid characters for sl no. '+ str(row[0]) +'\n'
                    #     write(logfile,propertyFile,sheet1.title,row[0],'Name has invalid characters',row[1])
                    #     #logfile.write(reason)
                    if config.INSERT_DATA  and pd.isna(row[29]):
                        validated = False
                        reason = 'Property File data sheet1 validation failed for sl no. '+ str(row[0]) + ', mobile no. is empty.\n'
                        write(logfile,propertyFile,sheet1.title,row[0],'mobile no. is empty',row[1])
                        #logfile.write(reason)
                    elif not pd.isna(row[29])  and  ( len(getMobileNumber(row[29],str,"")) != 10):
                        validated = False
                        count = count + 1
                        reason = 'Property File data sheet1 validation failed, Mobile number not correct for sl no. '+ str(row[0]) +'\n'
                        write(logfile,propertyFile,sheet1.title,row[0],'Mobile number not correct',row[1])
                        #logfile.write(reason)
                    # if not pd.isna(row[33]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[33]))):                        
                    #     validated = False
                    #     reason = 'Property File data validation failed, Guardian Name has invalid characters for abas id '+ str(row[0]) +'\n'
                    #     write(logfile,propertyFile,sheet1.title,None,'Guardian Name has invalid characters',row[0])
                    #     #logfile.write(reason)
                    if len(getValue(row[30], str, "")) > 0 and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[30]))) :                     
                        validated = False
                        reason = 'Property File data validation failed, Email id is not proper for abas id '+ str(row[1]) +'\n'
                        write(logfile,propertyFile,sheet1.title,row[0],'Email id is not proper',row[1])
                        #logfile.write(reason)
                    
                if pd.isna(row[7]):
                    validated = False
                    reason = 'Property File data validation failed for sl no. '+ str(row[0]) + ', usage category is empty.\n'
                    write(logfile,propertyFile,sheet1.title,row[0],'usage category is empty',row[1])
                    #logfile.write(reason)
                elif(str(row[7]).strip() == 'Commercial ( Nonresidential )' or str(row[7]).strip() == 'Institutional ( Nonresidential )'
                        or str(row[7]).strip() == 'Industrial ( Nonresidential )' or str(row[7]).strip() == 'Others ( Nonresidential )'):
                    if pd.isna(row[8]):
                        validated = False
                        reason = 'Property File data validation failed for sl no. '+ str(row[0]) + ', sub usage category is empty.\n'
                        write(logfile,propertyFile,sheet1.title,row[0],'sub usage category is empty',row[1])
                        #logfile.write(reason)
            except Exception as ex:
                print(config.CITY_NAME," validateDataForProperty Exception: ",ex)
                write(logfile,propertyFile,sheet1.title,row[0],str(ex) ,row[1])

        for index in range(3, sheet1.max_row +1):
            try: 
                if pd.isna(sheet1['B{0}'.format(index)].value):                    
                    break
                propSheetABASId = sheet1['B{0}'.format(index)].value
                if type(propSheetABASId) == int or type(propSheetABASId) == float:
                    propSheetABASId = str(int(sheet1['B{0}'.format(index)].value)) 
                abas_ids.append(str(propSheetABASId).strip())
            except Exception as ex:
                print( config.CITY_NAME,  " validateDataForProperty Exception: abas id is empty: ",ex)
        duplicate_ids = [item for item, count in collections.Counter(abas_ids).items() if count > 1]

        if(len(duplicate_ids) >= 1):
            validated = False
            reason = 'Property File data validation failed. ' +'Duplicate abas property id for '+ str(duplicate_ids) +'\n'
            # print(reason)
            write(logfile,propertyFile,sheet1.title,None,'Duplicate ABAS property id for '+ str(duplicate_ids))
            #logfile.write(reason)      
        
        sheet2_index =1
        for row in sheet2.iter_rows(min_row=2, max_col=12, max_row=sheet2.max_row ,values_only=True):
            try :
                sheet2_index = sheet2_index +1
                abasId = getValue(row[0], str, '')
                ownerName = getValue(row[2], str, '')
                if  len(abasId)==0  and len(ownerName) > 0   : 
                     write(logfile,propertyFile,sheet2.title,sheet2_index,'ABAS ID is empty ',None)
                validateMobile =False      
                if not pd.isna(row[0]):
                    if validateMobile and pd.isna(row[3]):
                        validated = False
                        reason = 'Property File data validation failed for abas id  '+ str(row[0]) + ', mobile no. is empty in multiple owner sheet.\n'
                        write(logfile,propertyFile,sheet2.title,None,'mobile no. is empty',row[0])
                        #logfile.write(reason)
                    elif not pd.isna(row[3]) and (len(str(row[3]).strip().replace(".0", "")) != 10):
                        validated = False
                        reason = 'Property File data validation failed, Mobile number not correct for abas id '+ str(row[0]) +'\n'
                        write(logfile,propertyFile,sheet2.title,None,'Mobile number not correct',row[0])
                        #logfile.write(reason)
                    if pd.isna(row[2]):
                        validated = False
                        reason = 'Property File data validation failed for abas id  '+ str(row[0]) + ', name is empty.\n'
                        write(logfile,propertyFile,sheet2.title,None,'name is empty',row[0])
                        #logfile.write(reason)
                    
                    # elif not pd.isna(row[2]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[2]))):                        
                    #     validated = False
                    #     reason = 'Property File data validation failed, Name has invalid characters for abas id '+ str(row[0]) +'\n'
                    #     write(logfile,propertyFile,sheet2.title,None,'Name has invalid characters',row[0])
                    #     #logfile.write(reason)
                    # if not pd.isna(row[7]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[7]))):                        
                    #     validated = False
                    #     reason = 'Property File data validation failed, Guardian Name has invalid characters for abas id '+ str(row[0]) +'\n'
                    #     write(logfile,propertyFile,sheet2.title,None,'Guardian Name has invalid characters',row[0])
                    #     #logfile.write(reason)
                    if not pd.isna(row[4]) and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[4]))):                        
                        validated = False
                        reason = 'Property File data validation failed, Email id is not proper for abas id '+ str(row[1]) +'\n'
                        write(logfile,propertyFile,sheet2.title,None,'Email id is not proper',row[0])
                        #logfile.write(reason)
            except Exception as ex:
                print(config.CITY_NAME," validateDataForProperty Exception: ",ex)
                write(logfile,propertyFile,sheet2.title,None,str(ex) ,row[0])
    except Exception as ex:
        print(config.CITY_NAME," validateDataForProperty Exception: ",ex)
         
    reason = 'Property file validation ends.\n'
    #print(reason)
    # logfile.write(reason)
    return validated

def ValidateCols(logfile, propertyFile, sheet1, sheet2):
    proper_column_order1 = ['Sl No.', 'Existing Property ID* ( Unique Value on which property are getting searched in existing system ) ', 
    'Old Property code*', 'Property Type *', 'Total Land Area (in sq feet)\u2009 *', 'Total constructed Area (in sq feet)*', 
    'Existing Usage Type', 'Usage type *', 'Sub Usage type *', 'Occupancy Type*', 'Number of Floors*', 'Number of Flats ', 'CB Name *', 
    'Locality / Mohalla *\u2009\n', 'Ward No.', 'Block No.', 'Street Name', 'Building / Colony Name', 'House / Door No', 'Pin Code', 
    'Location of the Property *', 'ARV', 'RV', 'Financial Year', 'Is Property on Dispute(Yes/No) ', 'Is Property Authorized(Yes/No) ', 
    'Is Property  Encroached(Yes/No) ', 'OwnerShip Type*', 'Name/ Authorized Person*', 'Mobile No.*', 'Email Id', 'Gender*', 'DOB', 'Guardian Name*', 
    'Relationship*', 'Is correspondence address, same As Property Address*', 'Correspondence Address*', 'Special category*', 'Institution Name', 
    'Institution Type', 'Designation', 'Landline No']

    proper_column_order2 =  ['Existing Property ID*', 'OwnerShip Type*', 'Name*', 'Mobile No.*', 'Email Id', 'Gender*', 'DOB', 'Guardian Name*', 
    'Relationship*', 'Is correspondence address, same As Property Adderss*', 'Correspondence Address*', 'Special category*']
    validated = True
    column_list = [c.value for c in next(sheet1.iter_rows(min_row=2, max_row=2))]
    
    for i in range(0, 42):
        if(proper_column_order1[i].strip() != column_list[i].strip()) :
            print(column_list[i])
            validated = False
            write(logfile,propertyFile,sheet1.title,None,'Column order / name is not correct',column_list[i])
            # break

    column_list = [c.value for c in next(sheet2.iter_rows(min_row=1, max_row=1))]

    for i in range(0, 12):
        if(proper_column_order2[i].strip() != column_list[i].strip()) :
            print(column_list[i])
            validated = False
            write(logfile,propertyFile,sheet2.title,None,'Column order / name is not correct',column_list[i])
            # break
    
    return validated                     
def enterDefaultMobileNo(propertyFile, tenantMapping, cityname, waterFile, sewerageFile, logfile):
    validated = True
    search_key = 'pb.'+ cityname
    res = list(tenantMapping.keys()).index(search_key)
    res = res+1
    res = res* 100000
    mobileNumber = 3000000000 + res + 0
    owner_obj = {}
    try:
        if os.path.exists(propertyFile) : 
            wb_property = openpyxl.load_workbook(propertyFile) 
            sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')   
            sheet2 = wb_property.get_sheet_by_name('Property Ownership Details')        
            for i in range(3, sheet1.max_row +1):
                #print('B{0}'.format(i))
                if pd.isna(sheet1['B{0}'.format(i)].value):                    
                    continue
                abas_id = sheet1['B{0}'.format(i)].value.strip()
                for row in sheet1.iter_rows(min_row=i, max_col=42, max_row=i,values_only=True):                    
                    owner = {}               
                    owner['mobileNumber'] = getMobileNumber(row[29],str,"")
                    owner['ownerType'] =  getValue(row[27],str,"") 
                    if abas_id not in owner_obj:
                        owner_obj[abas_id] = []
                    owner_obj[abas_id].append(owner)
            index = 2
            for row in sheet1.iter_rows(min_row=3, max_col=42, max_row=sheet1.max_row ,values_only=True):
                index = index + 1
                if pd.isna(row[1]):
                    continue 
                if (str(row[27]).strip() != 'Multiple Owners'):
                    if pd.isna(row[29]):
                        mobileNumber = mobileNumber + 1                    
                        value = 'AD{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
                        # logfile.write(value)
                        sheet1['AD{0}'.format(index)].value = mobileNumber
            index = 1
            for row in sheet2.iter_rows(min_row=2, max_col=5, max_row=sheet2.max_row ,values_only=True): 
                index = index + 1 
                if pd.isna(row[0]) :  
                    if (not pd.isna(row[1]) or not pd.isna(row[2])):
                        print("empty abas id in property file sheet2 while property validation")
                        write(logfile,propertyFile,sheet2.title,None,"empty abas id in property file sheet2 while property validation",None)
                        continue 
                    else : 
                        continue
                
                if pd.isna(row[3]):
                    mobileNumber = mobileNumber + 1
                    value = 'D{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
                    logfile.write(value)
                    # print(value)
                    sheet2['D{0}'.format(index)].value = mobileNumber
            wb_property.save(propertyFile)        
            wb_property.close()
        else:
            print("Property File doesnot exist for ", cityname)  

        # if os.path.exists(waterFile) :
        #     wb_water = openpyxl.load_workbook(waterFile) 
        #     water_sheet = wb_water.get_sheet_by_name('Water Connection Details') 
        #     index = 2
        #     for row in water_sheet.iter_rows(min_row=3, max_col=5, max_row=water_sheet.max_row ,values_only=True):
        #         index = index + 1
        #         if pd.isna(row[0]):
        #             continue
        #         if(str(row[3]).strip() == 'No'):
        #             if pd.isna(row[4]):
        #                 mobileNumber = mobileNumber + 1
        #                 value = 'E{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
        #                 logfile.write(value)
        #                 water_sheet['E{0}'.format(index)].value = mobileNumber
                
        #     wb_water.save(waterFile)        
        #     wb_water.close()
        #     wb_water = openpyxl.load_workbook(waterFile) 
        #     water_sheet = wb_water.get_sheet_by_name('Water Connection Details')
        #     for row in water_sheet.iter_rows(min_row=3, max_col=5, max_row=water_sheet.max_row ,values_only=True):
        #         if pd.isna(row[0]):
        #             continue
        #         if(str(row[3]).strip() == 'Yes'):
        #             for obj in owner_obj[str(row[1]).strip()]:
        #                 if(len(obj['mobileNumber']) == 0):
        #                     validated = False
        #                     reason = 'Mobile number in property is not available as in water template same as owner detail for abas id. '+ str(row[1]).strip() +'\n'
        #                     print(reason)
        #                     logfile.write(reason)
        # else:
        #     print("Water File doesnot exist for ", cityname)  

        # if os.path.exists(sewerageFile) :
        #     wb_sewerage = openpyxl.load_workbook(sewerageFile) 
        #     sewerage_sheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details') 
        #     index = 2
        #     for row in sewerage_sheet.iter_rows(min_row=3, max_col=5, max_row=sewerage_sheet.max_row ,values_only=True):
        #         index = index + 1
        #         if pd.isna(row[0]):
        #             continue
        #         if(str(row[3]).strip() == 'No'):
        #             if pd.isna(row[4]):
        #                 mobileNumber = mobileNumber + 1
        #                 value = 'E{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
        #                 logfile.write(value)
        #                 sewerage_sheet['E{0}'.format(index)].value = mobileNumber
                
        #     wb_sewerage.save(sewerageFile)        
        #     wb_sewerage.close()
        #     wb_sewerage = openpyxl.load_workbook(sewerageFile) 
        #     sewerage_sheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details') 
        #     for row in sewerage_sheet.iter_rows(min_row=3, max_col=10, max_row=sewerage_sheet.max_row ,values_only=True):
        #         if pd.isna(row[0]):
        #             continue
        #         if(str(row[3]).strip() == 'Yes'):
        #             for obj in owner_obj[str(row[1]).strip()]:
        #                 if(len(obj['mobileNumber']) == 0):
        #                     validated = False
        #                     reason = 'Mobile number in property is not available as in sewerage template same as owner detail for abas id. '+ str(row[1]).strip() +'\n'
        #                     logfile.write(reason)

        # else:
        #     print("Sewerage File doesnot exist for ", cityname) 
    except Exception as ex:
        print(config.CITY_NAME," DefaultMobileNo Exception: ",ex)
        traceback.print_exc()

    return validated

def createPropertyJson(sheet1, sheet2, locality_data,cityname, logfile,root, name):
    # abas_ids_multiple_owner = []
    # for index in range(2, sheet2.max_row +1):
    #     abas_ids_multiple_owner.append(sheet2['A{0}'.format(index)].value)
    createdCount = 0
    searchedCount = 0
    notCreatedCount = 0

    multiple_owner_obj = {}
    for i in range(2, sheet2.max_row +1):   
        try:
            if pd.isna(sheet2['A{0}'.format(i)].value):                    
                continue     
            abas_id = sheet2['A{0}'.format(i)].value.strip()
            for row in sheet2.iter_rows(min_row=i, max_col=12, max_row=i,values_only=True):                    
                owner = Owner()
                Owner.status = 'ACTIVE'
                owner.name = getValue( row[2] ,str,"NAME")
                owner.mobileNumber =  getMobileNumber( row[3] ,str,"3000000000")
                owner.emailId = getValue( row[4] ,str,"")
                owner.gender = process_gender( row[5] )
                if not pd.isna(row[6]):
                    owner.dob = getTime(row[6])
                owner.fatherOrHusbandName = getValue( row[7] ,str,"Guardian")
                owner.relationship =  process_relationship(str(row[8]).strip())
                owner.sameAsPeropertyAddress = getValue( row[9] ,str,"Yes")            
                owner.ownerType =  process_special_category(str(row[11]).strip())
                if abas_id not in multiple_owner_obj:
                    multiple_owner_obj[abas_id] = []
                multiple_owner_obj[abas_id].append(owner)
        except Exception as ex:
            print(config.CITY_NAME," createPropertyJson Exception: ",ex)
            traceback.print_exc()
    index = 2
    for row in sheet1.iter_rows(min_row=3, max_col=42, max_row=sheet1.max_row ,values_only=True):       
        try:   
            index = index + 1  
            property = Property()  
            property.abasPropertyId =  getValue( row[1] ,str,None)
            if pd.isna(property.abasPropertyId):
                print("empty Abas id in property file for sl no. ", row[0])
                continue     
            
            locality = Locality()
            address  = Address()
            owner = Owner()
            unit = Unit()
            constructionDetail = ConstructionDetail()
            institution = Institution()
            additionalDetail = AdditionalDetails()
            tenantId = 'pb.'+ cityname
            property.tenantId = tenantId
            
            property.oldPropertyId =  getValue( row[2] ,str,None)
            property.propertyType = process_property_type(str(row[3]).strip())
            if(row[4] == 0):
                row[4] = 1
            if(row[5] == 0):
                row[5] = 1
            property.landArea = getValue(row[4],float,1) 
            property.superBuiltUpArea = getValue(row[5],float,1) 
            property.usageCategory = process_usage_type(str(row[7]).strip())
            property.subUsageCategory = process_sub_usage_type(str(row[8]).strip())   
            if pd.isna(row[13]):
                locality.code = "LOCAL_OTHERS"   
            else:    
                locality.code = locality_data[row[13].strip()]
            address.city = cityname
            address.locality = locality
            address.location = process_location(str(row[20]).strip())
            address.street = getValue(row[16] ,str,None)
            address.buildingName = getValue(row[17] ,str,"")
            address.doorNo = getValue(row[18],str,"")
            address.pincode = getValue(row[19],str,None)
            correspondence_address = get_propertyaddress(address.doorNo,address.buildingName,getValue( row[13] ,str,"Others"),cityname)
            unit.occupancyType = process_occupancy_type(str(row[9]).strip())
            unit.arv = getValue(row[21],int,0) 
            unit.floorNo = 0
            if(len(property.subUsageCategory) == 0):
                unit.usageCategory = property.usageCategory
            else:                
                unit.usageCategory = property.subUsageCategory
            constructionDetail.builtUpArea = getValue(row[5],int,1) 
            unit.construction_detail = constructionDetail        
            property.address = address
            property.units = []
            property.units.append(unit)
            property.owners = []
            # converter = lambda  x,y  : x  if x is not pd.isna else y
            property.noOfFloors = getValue(row[10],int,1) 
            property.noOfFlats = getValue(row[11],int,0) 
            financial_year = getValue(row[23],str,"2020-2021").replace("-20", "-")
            property.financialYear = financial_year
            property.ownershipCategory = process_ownership_type(str(row[27]).strip())     
            if(property.ownershipCategory == 'INDIVIDUAL.SINGLEOWNER'):
                owner.status = 'ACTIVE'            
                owner.name = getValue(row[28] ,str,"NAME")
                owner.mobileNumber = getValue(row[29],str,"3000000000")
                owner.emailId = getValue(row[30] ,str,"")
                owner.gender = process_gender( row[31] )
                if not pd.isna(row[32]):
                    owner.dob = getTime(row[32])
                owner.fatherOrHusbandName = getValue(row[33],str,"Guardian")
                owner.relationship =  process_relationship(str(row[34]).strip())
                owner.sameAsPeropertyAddress = getValue(row[35],str,"Yes")
                if(owner.sameAsPeropertyAddress ==  'Yes'):
                    owner.correspondenceAddress = correspondence_address
                else: 
                    owner.correspondenceAddress = getValue(row[36],str,correspondence_address)
                owner.ownerType =  process_special_category(str(row[37]).strip())
                
                property.owners.append(owner)
            elif(property.ownershipCategory == 'INSTITUTIONALPRIVATE'):
                owner.status = 'ACTIVE'
                owner.name = getValue(row[28],str,"NAME")
                owner.mobileNumber = getValue(row[29],str,"3000000000")
                owner.emailId = getValue(row[30],str,"")
                owner.sameAsPeropertyAddress = getValue(row[35],str,"Yes")
                owner.correspondenceAddress = getValue(row[36],str,"")
                institution.name = getValue(row[38],str,"Institution")
                institution.type = process_private_institution_type(str(row[39]).strip())
                institution.designation = getValue(row[40],str,"Designation")
                owner.altContactNumber = getValue(row[41],str,"1000000000")
                if(owner.sameAsPeropertyAddress ==  'Yes'):
                    owner.correspondenceAddress = correspondence_address
                else: 
                    owner.correspondenceAddress = getValue(row[36],str,"Correspondence")
                owner.ownerType =  process_special_category(str(row[37]).strip())    
                property.institution = institution
                property.owners.append(owner)
            elif(property.ownershipCategory == 'INSTITUTIONALGOVERNMENT'):
                owner.status = 'ACTIVE'
                owner.name = getValue(row[28],str,"NAME")
                owner.mobileNumber = getValue(row[29],str,"3000000000")
                owner.emailId = getValue(row[30],str,"")
                owner.sameAsPeropertyAddress = getValue(row[35],str,"Yes")
                owner.correspondence_address = getValue(row[36],str,"")
                institution.name = getValue(row[38],str,"Institution")
                institution.type = process_govt_institution_type(str(row[39]).strip())
                institution.designation = getValue(row[40],str,"Designation")
                owner.altContactNumber = getValue(row[41],str,"1000000000")
                owner.sameAsPeropertyAddress = getValue(row[35],str,"Yes")
                if(owner.sameAsPeropertyAddress ==  'Yes'):
                    owner.correspondenceAddress = correspondence_address
                else: 
                    owner.correspondenceAddress = getValue(row[36],str,"Correspondence")
                owner.ownerType =  process_special_category(str(row[37]).strip())    
                property.institution = institution
                property.owners.append(owner)
            elif(property.ownershipCategory == 'INDIVIDUAL.MULTIPLEOWNERS'):
                for owner_obj in multiple_owner_obj[property.abasPropertyId]:
                    owner = owner_obj
                    owner.status = 'ACTIVE'  
                    if(owner.sameAsPeropertyAddress ==  'Yes'):
                        owner.correspondenceAddress = correspondence_address
                    else: 
                        owner.correspondenceAddress = getValue(row[36],str,"Correspondence")
                    property.owners.append(owner)

                # occurances = [i for i,x in enumerate(abas_ids_multiple_owner) if x == property.abas_property_id]
                # for i in occurances:                
                #     for row in sheet2.iter_rows(min_row=i+2, max_col=12, max_row=i+2,values_only=True):                    
                #         owner = Owner()
                #         owner.status = 'ACTIVE'
                #         owner.name = getValue(str(row[2]).strip(),str,"NAME")
                #         owner.mobileNumber = getValue(str(row[3]).strip(),str,"3000000001")
                #         owner.emailId = getValue(str(row[4]).strip(),str,"")
                #         owner.gender = process_gender(str(row[5]).strip())
                #         owner.fatherOrHusbandName = getValue(str(row[7]).strip(),str,"Guardian")
                #         owner.relationship =  process_relationship(str(row[8]).strip())
                #         owner.sameAsPeropertyAddress = getValue(str(row[9]).strip(),str,"Yes")
                #         if(owner.sameAsPeropertyAddress ==  'Yes'):
                #             owner.correspondenceAddress = correspondence_address
                #         else: 
                #             owner.correspondenceAddress = getValue(str(row[10]).strip(),str,"Correspondence")
                #         owner.ownerType =  process_special_category(str(row[11]).strip())
                #         property.owners.append(owner)

            additionalDetail.isRainwaterHarvesting = False
            additionalDetail.isPropertyDisputed = process_YesNo(str(row[24]))
            additionalDetail.isPropertyAuthorized = process_YesNo(str(row[25]))
            property.additional_details= additionalDetail
            property.source = 'LEGACY_RECORD'
            property.channel = 'MIGRATION'
            property.creationReason = 'DATA_UPLOAD'
        except Exception as ex:
            print(config.CITY_NAME," createPropertyJson Exception: ",ex)
            traceback.print_exc()
        auth_token = superuser_login()["access_token"]
        status, res = property.search_abas_property(auth_token, tenantId, property.abasPropertyId)
        with io.open(os.path.join(root, name,"property_search_res.json"), mode="w", encoding="utf-8") as f:
            json.dump(res, f, indent=2,  ensure_ascii=False)
        if(len(res['Properties']) == 0):
            
            statusCode, res = property.upload_property(auth_token, tenantId, property.abasPropertyId,root, name,)
            with io.open(os.path.join(root, name,"property_create_res.json"), mode="w", encoding="utf-8") as f:
                json.dump(res, f, indent=2,  ensure_ascii=False)
            propertyId = ''
            if(statusCode == 200 or statusCode == 201):
                for found_index, resProperty in enumerate(res["Properties"]):
                    propertyId = resProperty["propertyId"]
                    # value = 'B{0}'.format(index) + '    ' + str(propertyId) + '\n'
                    # logfile.write(value)
                    sheet1['AQ{0}'.format(index)].value = propertyId
                    reason = 'property created for abas id ' + str(property.abasPropertyId)
                    # logfile.write(reason)
                    # print(reason)
                    createdCount = createdCount + 1
                    break
            else:
                reason = 'property not created status code '+ str(statusCode) + ' for abas id ' + str(property.abasPropertyId) + ' response: ', str(res)  + '\n'
                # logfile.write(reason)
                print(reason)
                notCreatedCount = notCreatedCount + 1
        else:
            for found_index, resProperty in enumerate(res["Properties"]):
                propertyId = resProperty["propertyId"]
                break
            sheet1['AQ{0}'.format(index)].value = propertyId
            reason = 'property already exist for abas id '+ str(property.abasPropertyId) + '\n'
            # logfile.write(reason)
            # print(reason)
            searchedCount = searchedCount + 1

    reason = 'property created count: '+ str(createdCount)
    print(reason)
    reason = 'property not created count: '+ str(notCreatedCount)
    print(reason)
    reason = 'property searched count: '+ str(searchedCount)
    print(reason)


def get_propertyaddress(doorNo, buildingName,locality,cityname):
    return doorNo + ' ' + buildingName + ' ' +locality + ' ' + cityname

def process_YesNo(value):
    YesNo_MAP = {
        "Yes": True,
        "No": False,
        "None": None
    }
    return YesNo_MAP[value]

def process_location(value):
    location_MAP = {
        "Inside Civil Area (Bazar Area)": "CIVIL",
        "Outside Civil Area (Bungalow Area)": "BUNGLOW",
        "Outside Cantonment Area": "OUTSIDE",
        "None": "CIVIL"
    }
    return location_MAP[value]

def process_relationship(value):
    if value is None : 
        value ="parent"
    value = value.strip().lower()
    relationship_MAP = {
        "parent": "PARENT",
        "spouse": "SPOUSE",
        "gurdian": "GUARDIAN",
        "none": "PARENT"
    }
    return relationship_MAP[value]

def process_gender(value):
    if value is None : 
        value ="Male"
    value =str(value).strip().lower()
    gender_MAP = {
        "male": "MALE",
        "female": "FEMALE",
        "transgender": "TRANSGENDER",
        "none": "MALE"       
    }
    return gender_MAP[value]

def process_private_institution_type(value):
    private_institution_MAP = {
        "None": "OTHERSPRIVATEINSTITUITION",
        "Private company": "PRIVATECOMPANY",
        "NGO": "NGO",
        "Private Trust": "PRIVATETRUST",
        "Private Board": "PRIVATEBOARD",
        "Other Private Institution": "OTHERSPRIVATEINSTITUITION"        
    }
    return private_institution_MAP[value]

def process_govt_institution_type(value):
    govt_institution_MAP = {
        "None": "OTHERGOVERNMENTINSTITUITION",
        "CB Government": "ULBGOVERNMENT",
        "state Government": "INSTITUTIONALGOVERNMENT",
        "Central Government": "CENTRALGOVERNMENT",
        "Other Government Institution": "OTHERGOVERNMENTINSTITUITION"
    }
    return govt_institution_MAP[value]

def process_property_type(value):
    PT_MAP = {
        "None": "BUILTUP.INDEPENDENTPROPERTY",
        "Vacant Land": "VACANT",
        "Flat/Part of the building": "BUILTUP.SHAREDPROPERTY",
        "Independent Building": "BUILTUP.INDEPENDENTPROPERTY"
    }
    return PT_MAP[value]

def process_occupancy_type(value):
    OC_MAP = {
        "None": "SELFOCCUPIED",
        "Self-Occupied": "SELFOCCUPIED",
        "Rented": "RENTED",
        "Unoccupied": "UNOCCUPIED"
    }
    return OC_MAP[value]

def process_usage_type(value):
    USAGE_MAP = {
        "Residential": "RESIDENTIAL",
        "Nonresidential ( Nonresidential )": "NONRESIDENTIAL.NONRESIDENTIAL",
        "Commercial ( Nonresidential )": "NONRESIDENTIAL.COMMERCIAL",
        "Industrial ( Nonresidential )": "NONRESIDENTIAL.INDUSTRIAL",
        "Institutional ( Nonresidential )": "NONRESIDENTIAL.INSTITUTIONAL",
        "Others ( Nonresidential )": "NONRESIDENTIAL.OTHERS",
        "Mixed": "MIXED",
        "Slum": "SLUM",
        "None": "RESIDENTIAL"
    }
    return USAGE_MAP[value]

def process_sub_usage_type(value):  
    SUB_USAGE_MAP = {
        'None':'' ,'Animal Dairy(Below 10 Cattle)':'NONRESIDENTIAL.COMMERCIAL.ANIMALDAIRYLESS' , 'Animal Dairy(Above 10 Cattle)':'NONRESIDENTIAL.COMMERCIAL.ANIMALDAIRYMORE' , 'Bank':'NONRESIDENTIAL.COMMERCIAL.BANK' , 'Dhobi':'NONRESIDENTIAL.COMMERCIAL.DHOBI' , 'Dyers':'NONRESIDENTIAL.COMMERCIAL.DYERS' , 'Movie Theatre':'NONRESIDENTIAL.COMMERCIAL.ENTERTAINMENT.MOVIETHEATRE' , 'Multiplex':'NONRESIDENTIAL.COMMERCIAL.ENTERTAINMENT.MULTIPLEX' , 'Marriage Palace':'NONRESIDENTIAL.COMMERCIAL.EVENTSPACE.MARRIAGEPALACE' , 'Ac Restaurant':'NONRESIDENTIAL.COMMERCIAL.FOODJOINTS.ACRESTAURANT' , 'Non Ac Restaurant':'NONRESIDENTIAL.COMMERCIAL.FOODJOINTS.NONACRESTAURANT' , 'Bhojanalaya/Tea Shop/Halwai Shop':'NONRESIDENTIAL.COMMERCIAL.FOODJOINTS.TEA' , 'Hotels':'NONRESIDENTIAL.COMMERCIAL.HOTELS' , 'Pathlab':'NONRESIDENTIAL.COMMERCIAL.MEDICALFACILITY.PATHLAB' , 'Private Dispensary':'NONRESIDENTIAL.COMMERCIAL.MEDICALFACILITY.PVTDISPENSARY' , 'Private Hospital':'NONRESIDENTIAL.COMMERCIAL.MEDICALFACILITY.PVTHOSPITAL' , 'Office Space(Less Than 10 Persons)':'NONRESIDENTIAL.COMMERCIAL.OFFICESPACELESS' , 'Office Space(More Than 10 Persons)':'NONRESIDENTIAL.COMMERCIAL.OFFICESPACEMORE' , 'Other Commercial Usage':'NONRESIDENTIAL.COMMERCIAL.OTHERCOMMERCIALSUBMINOR.OTHERCOMMERCIAL' , 'Petrol Pump':'NONRESIDENTIAL.COMMERCIAL.PETROLPUMP' , 'Grocery Store':'NONRESIDENTIAL.COMMERCIAL.RETAIL.GROCERY' , 'Malls':'NONRESIDENTIAL.COMMERCIAL.RETAIL.MALLS' , 'Pharmacy':'NONRESIDENTIAL.COMMERCIAL.RETAIL.PHARMACY' , 'Showroom':'NONRESIDENTIAL.COMMERCIAL.RETAIL.SHOWROOM' , 'Service Centre':'NONRESIDENTIAL.COMMERCIAL.SERVICECENTER' , 'Statutory Organisation':'NONRESIDENTIAL.COMMERCIAL.STATUTORY.STATUTORYORGANISATION' , 'Manufacturing Facility(Less Than 10 Persons)':'NONRESIDENTIAL.INDUSTRIAL.MANUFACTURINGFACILITY.MANUFACTURINGFACILITYLESS' , 'Manufacturing Facility(More Than 10 Persons)':'NONRESIDENTIAL.INDUSTRIAL.MANUFACTURINGFACILITY.MANUFACTURINGFACILITYMORE' , 'Other Industrial Usage':'NONRESIDENTIAL.INDUSTRIAL.OTHERINDUSTRIALSUBMINOR.OTHERINDUSTRIAL' , 'Godown/Warehouse':'NONRESIDENTIAL.INDUSTRIAL.WAREHOUSE.WAREHOUSE' , 'College':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.COLLEGES' , 'Other Private Educational Institute':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.OTHEREDUCATIONAL' , 'Polytechnic':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.POLYTECHNICS' , 'School':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.SCHOOL' , 'Training Institute':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.TRAININGINSTITUTES' , 'Govt. Aided Educational Institute':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONALGOVAIDED.GOVAIDEDEDUCATIONAL' , 'Historical Building':'NONRESIDENTIAL.INSTITUTIONAL.HISTORICAL.HISTORICAL' , 'Stray Animal Care Center':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.ANIMALCARE' , 'Home For The Disabled / Destitute':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.DISABLEDHOME' , 'Old Age Homes':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.OLDAGEHOMES' , 'Orphanage':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.ORPHANAGE' , 'Others':'NONRESIDENTIAL.INSTITUTIONAL.OTHERINSTITUTIONALSUBMINOR.OTHERINSTITUTIONAL' , 'Community Hall':'NONRESIDENTIAL.INSTITUTIONAL.PUBLICFACILITY.COMMUNITYHALL' , 'Govt. Hospital & Dispensary':'NONRESIDENTIAL.INSTITUTIONAL.PUBLICFACILITY.GOVTHOSPITAL' , 'Public Libraries':'NONRESIDENTIAL.INSTITUTIONAL.PUBLICFACILITY.LIBRARIES' , 'Golf Club':'NONRESIDENTIAL.INSTITUTIONAL.RECREATIONAL.GOLFCLUB' , 'Social Club':'NONRESIDENTIAL.INSTITUTIONAL.RECREATIONAL.SOCIALCLUB' , 'Sports Stadium':'NONRESIDENTIAL.INSTITUTIONAL.RECREATIONAL.SPORTSSTADIUM' , 'Religious':'NONRESIDENTIAL.INSTITUTIONAL.RELIGIOUSINSTITUITION.RELIGIOUS' , 'Cremation/ Burial Ground':'NONRESIDENTIAL.OTHERS.PUBLICFACILITY.CREMATIONBURIAL'
    }
    return SUB_USAGE_MAP[value]

def process_ownership_type(value):
    if value is None : 
        value ="None"
    value =value.strip().lower()
    Ownsership_MAP = {
        "none": "INDIVIDUAL.SINGLEOWNER",
        "single owner": "INDIVIDUAL.SINGLEOWNER",
        "multiple owners": "INDIVIDUAL.MULTIPLEOWNERS",
        "institutional- private": "INSTITUTIONALPRIVATE",
        "institutional- government": "INSTITUTIONALGOVERNMENT"
    }
    return Ownsership_MAP[value]

def process_special_category(value):
    if value is None: 
        value ="None"
    value =value.strip().lower()
    special_category_MAP = {
        "freedom fighter": "FREEDOMFIGHTER",
        "widow": "WIDOW",
        "handicapped": "HANDICAPPED",
        "below poverty line": "BPL",
        "defense personnel": "DEFENSE",
        "employee/staff of cb": "STAFF",
        "none of the above": "NONE",
        "none":"NONE"
    }
    return special_category_MAP[value]

# def getValue(value,dataType,defValue="") :
#     try:
#         if(value == None or value == 'None' or pd.isna(value)): 
#             return defValue    
#         else : 
#             if dataType ==str : 
#                 return dataType(value).strip()
#             elif dataType == float:
#                 return round(value, 2)
#             else : 
#                 return dataType(value)
#     except: 
#         return defValue

def getMobileNumber(value,dataType,defValue="") :
    if(value == None or value == 'None' or pd.isna(value)): 
        return defValue    
    else : 
        if type(value) == int or type(value) == float:
            value = str(int(value)) 
        return dataType(value).strip()
        

if __name__ == "__main__":
    main()    
              


