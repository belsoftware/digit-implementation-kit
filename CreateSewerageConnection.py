from common import *
from config import config
import io
import os 
import sys
from SewerageConnection import *
from PropertyTax import *
import pandas as pd
# from CreateProperty import getValue
import openpyxl
import collections
import traceback

def main():
    Flag =False
    
def ProcessSewerageConnection(propertyFile, sewerageFile, logfile, root, name,  cityname) :
    wb_property = openpyxl.load_workbook(propertyFile) 
    propertySheet = wb_property.get_sheet_by_name('Property Assembly Detail') 
    wb_sewerage = openpyxl.load_workbook(sewerageFile) 
    sewerageSheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details')  
    #print('no. of rows in sewerage file: ', sewerageSheet.max_row +1 ) 
    validate = validateSewerageData(propertySheet, sewerageFile, logfile, cityname)  
    if(validate == False):                
        print('Data validation for sewerage Failed, Please check the log file.') 
        if config.INSERT_DATA: 
            return
    else:
        print('Data validation for sewerage success.')
    if config.INSERT_DATA: 
        createSewerageJson(propertySheet, sewerageSheet, cityname, logfile, root, name)   
        wb_sewerage.save(sewerageFile)        
        wb_sewerage.close()

def validateSewerageData(propertySheet, sewerageFile, logfile, cityname):
    validated = True
    wb_sewerage = openpyxl.load_workbook(sewerageFile) 
    sewerage_sheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details') 
    index = 2
    reason = 'sewerage file validation starts.\n'
    #print(reason)
    #logfile.write(reason)
    abas_ids = [] 
    old_connections = []
    try:
        for index in range(3, propertySheet.max_row +1 ):
            if pd.isna(propertySheet['B{0}'.format(index)].value):
                # validated = False
                # reason = 'Sewerage File data validation failed, Sl no. column is empty'
                # #logfile.write(reason)
                # write(logfile,"property excel",propertySheet.title,index,'Sl no. column is empty')
                continue
            propSheetABASId = propertySheet['B{0}'.format(index)].value
            if type(propSheetABASId) == int or type(propSheetABASId) == float:
                propSheetABASId = str(int(propertySheet['B{0}'.format(index)].value)) 
            abas_ids.append(str(propSheetABASId).strip())   
    except Exception as ex:
        print(config.CITY_NAME," validateSewerageData Exception: ", ex)
    emptyRows =0 
    for row in sewerage_sheet.iter_rows(min_row=3, max_col=22, max_row=sewerage_sheet.max_row +1 ,values_only=True):        
        index = index + 1
        try:
            if emptyRows > 10 :
                break
            if pd.isna(row[1]):
                emptyRows = emptyRows +1
                continue
            if pd.isna(row[0]):
                validated = False
                reason = 'Sewerage File data validation failed, Sl no. column is empty\n'
                #logfile.write(reason)
                write(logfile,sewerageFile,sewerage_sheet.title,row[0],'Sl no. column is empty',row[1])
            if pd.isna(row[1]):
                validated = False
                reason = 'Sewerage File data validation failed for sl no. '+ str(row[0]) + ', abas id is empty.\n'
                #logfile.write(reason) 
                write(logfile,sewerageFile,sewerage_sheet.title,row[0],'abas id is empty',row[1])
            if pd.isna(row[2]):
                validated = False
                reason = 'Sewerage File data validation failed for sl no. '+ str(row[0]) + ', old connection number is empty.\n'
                #logfile.write(reason)
                write(logfile,sewerageFile,sewerage_sheet.title,row[0],'old connection number is empty',row[1])

            if pd.isna(row[3]):
                validated = False
                reason = 'Sewerage File data validation failed for sl no. '+ str(row[0]) + ', same as property address cell is empty.\n'
                #logfile.write(reason)
                write(logfile,sewerageFile,sewerage_sheet.title,row[0],'same as property address cell is empty',row[1])
            if(str(row[3]).strip() == 'No'):
                if pd.isna(row[4]) :
                    validated = False
                    reason = 'Sewerage File data validation failed for sl no. '+ str(row[0]) + ', mobile number is empty.\n'
                    #logfile.write(reason) 
                    write(logfile,sewerageFile,sewerage_sheet.title,row[0],'mobile number is empty',row[1])
                elif not pd.isna(row[4]) and (len(str(row[4]).strip().replace(".0", "")) != 10):
                        validated = False
                        reason = 'Sewerage File data validation failed, Mobile number not correct for abas id '+ str(row[0]) +'\n'
                        write(logfile,sewerageFile,sewerage_sheet.title,None,'Mobile number not correct',row[0])
                        #logfile.write(reason)
                if pd.isna(row[5]):
                    validated = False
                    reason = 'Sewerage File data validation failed for sl no. '+ str(row[0]) + ',name is empty.\n'
                    #logfile.write(reason) 
                    write(logfile,sewerageFile,sewerage_sheet.title,row[0],' name is empty',row[1])
                # elif not pd.isna(row[5]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[5]))):
                #     validated = False
                #     reason = 'Sewerage File data validation failed, Name has invalid characters for sl no. '+ str(row[0]) +'\n'
                #     #logfile.write(reason)  
                #     write(logfile,sewerageFile,sewerage_sheet.title,row[0],'Name has invalid characters',row[1])
                # if not pd.isna(row[9]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[9]))):                        
                #     validated = False
                #     reason = 'Sewerage File data validation failed, Guardian Name has invalid characters for abas id '+ str(row[0]) +'\n'
                #     write(logfile,sewerageFile,sewerage_sheet.title,None,'Guardian Name has invalid characters',row[0])
                #     #logfile.write(reason)
                if len(getValue(row[6], str, "")) > 0 and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[6]))) :                      
                    validated = False
                    reason = 'Sewerage File data validation failed, Email id is not proper for abas id '+ str(row[1]) +'\n'
                    write(logfile,sewerageFile,sewerage_sheet.title,row[0],'Email id is not proper',row[1])
                    #logfile.write(reason)
            if not pd.isna(row[1]):
                abasid = row[1]
                if type(abasid) == int or type(abasid) ==float : 
                    abasid = str(int (abasid))
                if str(abasid).strip() not in abas_ids:
                    validated = False
                    reason = 'there is no abas id available in property data for sewerage connection sl no. '+ str(row[0]) +'\n'
                    #logfile.write(reason) 
                    write(logfile,sewerageFile,sewerage_sheet.title,row[0],'ABAS id not available in property data',row[1])
        except Exception as ex:
            # print(config.CITY_NAME," validateSewerageData Exception: ", row[0], '   ', ex)
            write(logfile,sewerageFile,sewerage_sheet.title,row[0],str(ex) ,row[1])
    for index in range(3, sewerage_sheet.max_row +1):
        try:
            if pd.isna(sewerage_sheet['B{0}'.format(index)].value):                    
                break
            oldConnectionNo = sewerage_sheet['C{0}'.format(index)].value
            if type(oldConnectionNo) == int or type(oldConnectionNo) == float:
                oldConnectionNo = str(int(sewerage_sheet['C{0}'.format(index)].value)) 
            old_connections.append(str(oldConnectionNo).strip())
        except Exception as ex:
            print( config.CITY_NAME,  " validateDataForSewerage Exception: existing sewerage connection no is empty: ",ex)
            traceback.print_exc()
            write(logfile,sewerageFile,sewerage_sheet.title,row[0],'existing sewerage connection no is empty',row[1])
    duplicate_ids = [item for item, count in collections.Counter(old_connections).items() if count > 1]

    if(len(duplicate_ids) >= 1):
        validated = False
        reason = 'Sewerage File data validation failed. ' +'Duplicate existing sewerage connection no for '+ str(duplicate_ids) +'\n'
        # print(reason)
        write(logfile,sewerageFile,sewerage_sheet.title,None,'Duplicate existing sewerage connection no for '+ str(duplicate_ids))
        #logfile.write(reason)  
    reason = 'sewerage file validation ends.\n'
    #print(reason)
    #logfile.write(reason) 
    return validated


def createSewerageJson(propertySheet, sewerageSheet, cityname, logfile, root, name):
    createdCount = 0
    searchedCount = 0
    notCreatedCount = 0
    owner_obj = {}
    for i in range(3, propertySheet.max_row +1 ):    
        try:    
            abas_id = propertySheet['B{0}'.format(i)].value.strip()
            for row in propertySheet.iter_rows(min_row=i, max_col=42, max_row=i,values_only=True):                    
                owner = Owner()
                address = Address()
                address.buildingName = getValue(row[17].strip(),str,"")
                address.doorNo = getValue(str(row[18]).strip(),str,"")
                correspondence_address = get_propertyaddress(address.doorNo,address.buildingName,getValue(str(row[13]).strip(),str,"Others"),cityname)
                owner.name = getValue(str(row[28]).strip(),str,"NAME")
                owner.mobileNumber = getValue(str(row[29]).strip(),str,"3000000000")
                owner.emailId = getValue(str(row[30]).strip(),str,"")
                owner.gender = process_gender(str(row[31]).strip())
                owner.fatherOrHusbandName = getValue(str(row[33]).strip(),str,"Guardian")
                owner.relationship =  process_relationship(str(row[34]).strip())
                owner.sameAsPeropertyAddress = getValue(str(row[35]).strip(),str,"Yes")
                if(owner.sameAsPeropertyAddress ==  'Yes'):
                    owner.correspondenceAddress = correspondence_address
                else: 
                    owner.correspondenceAddress = getValue(str(row[36]).strip(),str,correspondence_address)
                owner.ownerType =  process_special_category(str(row[37]).strip())
                if abas_id not in owner_obj:
                    owner_obj[abas_id] = []
                owner_obj[abas_id].append(owner)
        except Exception as ex:
            print(config.CITY_NAME," createSewerageJson Exception: ", row[0], '   ', ex)
    index = 2
    for row in sewerageSheet.iter_rows(min_row=3, max_col=19, max_row=sewerageSheet.max_row +1 , values_only=True):
        try:  
            index = index + 1
            abasPropertyId =  getValue(str(row[1]).strip(),str,None)  
            property = Property() 
            auth_token = superuser_login()["access_token"]
            tenantId = 'pb.'+ cityname
            property.tenantId = tenantId
            if pd.isna(abasPropertyId):
                print("empty Abas id in sewerage file for sl no. ", row[0])
                break
            
            status, res = property.search_abas_property(auth_token, tenantId, abasPropertyId)        
            with io.open(os.path.join(root, name,"property_search_res.json"), mode="w", encoding="utf-8") as f:
                json.dump(res, f, indent=2,  ensure_ascii=False) 
            propertyId = ''
            if(len(res['Properties']) >= 1):
                for found_index, resProperty in enumerate(res["Properties"]):
                    propertyId = resProperty["propertyId"]
                    break
                sewerageConnection = SewerageConnection()
                connectionHolder = ConnectionHolder()
                processInstance = ProcessInstance()
                additionalDetail = AdditionalDetail()
                sewerageConnection.connectionHolders = []
                if(str(row[3]).strip() == 'Yes'):
                    sewerageConnection.connectionHolders = None
                else:
                    connectionHolder.name = getValue(str(row[5]).strip(),str,"NAME")
                    connectionHolder.mobileNumber = getValue(str(row[4]).strip(),str,"3000000000")
                    connectionHolder.fatherOrHusbandName = getValue(str(row[9]).strip(),str,"Guardian")
                    connectionHolder.emailId = getValue(str(row[6]).strip(),str,"")
                    connectionHolder.correspondenceAddress = getValue(str(row[12]).strip(),str,"Correspondence")
                    connectionHolder.relationship = process_relationship(str(row[10]).strip())
                    connectionHolder.ownerType = process_special_category(str(row[13]).strip())
                    connectionHolder.gender = process_gender(str(row[7]).strip())
                    connectionHolder.sameAsPropertyAddress = False
                    sewerageConnection.connectionHolders.append(connectionHolder)
                
                sewerageConnection.oldConnectionNo = getValue(str(row[2]).strip(),str,None)
                sewerageConnection.drainageSize = getValue(str(row[14]).strip(),float,0.25)
                sewerageConnection.proposedDrainageSize = getValue(str(row[14]).strip(),float,0.25)
                if(sewerageConnection.waterSource != 'OTHERS'):
                    sewerageConnection.waterSubSource = sewerageConnection.waterSource.split('.')[1]                
                else:
                    sewerageConnection.waterSubSource = ''
                    sewerageConnection.sourceInfo = 'Other'
                sewerageConnection.propertyOwnership  = process_propertyOwnership(str(row[11]))
                sewerageConnection.noOfWaterClosets = getValue(str(row[15]).strip(),int,1)
                sewerageConnection.proposedWaterClosets = getValue(str(row[15]).strip(),int,1)
                sewerageConnection.noOfToilets = getValue(str(row[16]).strip(),int,1)
                sewerageConnection.proposedToilets = getValue(str(row[16]).strip(),int,1)
                if not pd.isna(row[17]):
                    waterConnection.connectionExecutionDate = getTime(row[17])
                additionalDetail.locality = ''
                sewerageConnection.additionalDetails = additionalDetail
                processInstance.action = 'ACTIVATE_CONNECTION'
                sewerageConnection.tenantId = tenantId
                sewerageConnection.propertyId = propertyId
                sewerageConnection.processInstance = processInstance
                sewerageConnection.water = False
                sewerageConnection.sewerage = True
                sewerageConnection.service = 'Sewerage'
                sewerageConnection.applicationType = 'NEW_SEWERAGE_CONNECTION'
                sewerageConnection.applicationStatus = 'CONNECTION_ACTIVATED'
                sewerageConnection.source = 'MUNICIPAL_RECORDS'
                sewerageConnection.channel = 'DATA_ENTRY'
                sewerageConnection.status = 'ACTIVE'
        except Exception as ex:
            print("createSewerageJson Exception: ", row[0], '   ', ex)

            auth_token = superuser_login()["access_token"]
            status, res = sewerageConnection.search_sewerage_connection(auth_token, tenantId, sewerageConnection.oldConnectionNo)               
            with io.open(os.path.join(root, name,"sewerage_search_res.json"), mode="w", encoding="utf-8") as f:
                json.dump(res, f, indent=2,  ensure_ascii=False)  
            if(len(res['SewerageConnections']) == 0):
                statusCode, res = sewerageConnection.upload_sewerage(auth_token, tenantId, sewerageConnection.oldConnectionNo, root, name)
                with io.open(os.path.join(root, name,"sewerage_create_res.json"), mode="w", encoding="utf-8") as f:
                    json.dump(res, f, indent=2,  ensure_ascii=False)  
                sewerageconnectionNo = '' 
                if(statusCode == 200 or statusCode == 201):
                    for found_index, resProperty in enumerate(res["SewerageConnections"]):
                        connectionNo = resProperty["connectionNo"]
                        value = 'B{0}'.format(index) + '    ' + str(connectionNo) + '\n'
                        logfile.write(value)
                        sewerageSheet['T{0}'.format(index)].value = connectionNo
                        reason = 'sewerage connection created for abas id ' + str(property.abasPropertyId)
                        # logfile.write(reason)
                        # print(reason)
                        createdCount = createdCount + 1
                else:
                    reason = 'sewerage not created status code '+ str(statusCode)+ ' for abas id ' + str(property.abasPropertyId) + ' response: '+ str(res) + '\n'
                    # logfile.write(reason)
                    print(reason)
                    notCreatedCount = notCreatedCount + 1
            else:
                reason = 'sewerage connection exist for abas id ' + str(property.abasPropertyId)
                # logfile.write(reason)
                # print(reason)
                searchedCount = searchedCount + 1
                        
        else:
            reason = 'property does not exist for abas id '+ str(property.abasPropertyId) + '\n'
            # logfile.write(reason)
            

    reason = 'sewerage created count: '+ str(createdCount)
    print(reason)
    reason = 'sewerage not created count: '+ str(notCreatedCount)
    print(reason)
    reason = 'sewerage searched count: '+ str(searchedCount)
    print(reason)


def getTime(dateObj,defValue=None) :
    try : 
        if not isinstance(dateObj, datetime) and type(dateObj) is str: 
            dateStr =dateObj.strip() 
            if "/" in dateStr : 
                dateObj=datetime.strptime(dateStr, '%d/%m/%Y') 
            elif "." in dateStr : 
                dateObj=datetime.strptime(dateStr, '%d.%m.%Y') 
            else : 
                dateObj=datetime.strptime(dateStr, '%d-%m-%Y') 
        milliseconds = int((dateObj - datetime(1970, 1, 1)).total_seconds())*1000
        return milliseconds
    except Exception as ex:
        print("Error in time conversion ",dateObj,ex)
    return None

def get_propertyaddress(doorNo, buildingName,locality,cityname):
    return doorNo + ' ' + buildingName + ' ' +locality + ' ' + cityname


def process_relationship(value):
    if value is None : 
        value ="parent"
    value = value.strip().lower()
    relationship_MAP = {
        "parent": "PARENT",
        "spouse": "SPOUSE",
        "gurdian": "GUARDIAN",
        "none": "PARENT"
    }
    return relationship_MAP[value]

def process_gender(value):
    if value is None : 
        value ="Male"
    value =str(value).strip().lower()
    gender_MAP = {
        "male": "MALE",
        "female": "FEMALE",
        "transgender": "TRANSGENDER",
        "none": "MALE"       
    }
    return gender_MAP[value]

def process_connection_type(value):
    connection_MAP = {
        "Metered": "Metered",
        "Non-Metered": "Non Metered",
        "None": "Non Metered"     
    }
    return connection_MAP[value]

def process_propertyOwnership(value):
    propertyOwnership_MAP = {
        "None": None,
        "HOR": "HOR",
        "Tenant": "TENANT"
    }
    return propertyOwnership_MAP[value]

def process_connection_permission(value):
    connection_permission_MAP = {
        "Authorized": "AUTHORIZED",
        "Unauthorized": "UNAUTHORIZED",
        "None": "AUTHORIZED"
    }
    return connection_permission_MAP[value]

def process_special_category(value):
    special_category_MAP = {
        "Freedom fighter": "FREEDOMFIGHTER",
        "Widow": "WIDOW",
        "Handicapped": "HANDICAPPED",
        "Below Poverty Line": "BPL",
        "Defense Personnel": "DEFENSE",
        "Employee/Staff of CB": "STAFF",
        "None of the above": "NONE",
        "None":"NONE"
    }
    return special_category_MAP[value]

def getValue(value,dataType,defValue="") :
    if(value == None or value == 'None' or pd.isna(value)): 
        return defValue    
    
    else : 
        if dataType ==str : 
            return dataType(value).strip()
        else : 
            return dataType(value)

if __name__ == "__main__":
    main()    
    type         


